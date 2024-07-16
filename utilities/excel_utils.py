import openpyxl
from openpyxl.styles import PatternFill


def get_row_count(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.max_row


def get_column_count(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.max_column


def read_data(file, sheetname, row_num, column_num):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.cell(row_num, column_num).value


def write_data(file, sheetname, row_num, column_num, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    sheet.cell(row_num, column_num).value = data
    workbook.save(file)


def fill_green(file, sheetname, row_num, column_num):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    greenfill = PatterFill(start_colour='00FF00', end_colour='00FF00', patternType='solid')
    sheet.cell(row_num, column_num).fill = greenfill
    workbook.save(file)


def fill_red(file, sheetname, row_num, column_num):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    redfill = PatterFill(start_colour='FF0000', end_colour='FF0000', patternType='solid')
    sheet.cell(row_num, column_num).fill = redfill
    workbook.save(file)